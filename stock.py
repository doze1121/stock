import random
import time
import openpyxl

from selenium import webdriver
from selenium_stealth import stealth
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

options = webdriver.ChromeOptions()
options.add_argument("start-maximized")
options.add_argument('--blink-settings=imagesEnabled=false')

options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option('useAutomationExtension', False)
options.page_load_strategy = 'normal'

driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 3)

stealth(driver,
    user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36',
    languages=["en-US", "en"],
    vendor="Google Inc.",
    platform="Win32",
    webgl_vendor="Intel Inc.",
    renderer="Intel Iris OpenGL Engine",
    fix_hairline=False,
    run_on_insecure_origins=False,
    )

workbook = openpyxl.load_workbook('/Users/User/Desktop/Кросы/chrome-win64/in.xlsx')
worksheet = workbook.active

sizes = [header.value for header in worksheet[1][12:]]

try:
    for row, link in enumerate(worksheet['D'], 1):
        if row <= 1:
            continue
        driver.get(link.value)
        driver.implicitly_wait(10)

        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#menu-button-pdp-size-selector')))
        driver.find_element(By.CSS_SELECTOR, '#menu-button-pdp-size-selector').click()
        flag = False
        try:
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#menu-list-pdp-size-selector > div.css-0 > div')))
            sizes_bar = driver.find_element(By.CSS_SELECTOR, '#menu-list-pdp-size-selector > div.css-0 > div')
            sizes_buttons = sizes_bar.find_elements(By.CSS_SELECTOR, '[data-testid="size-conversion-chip"]')
            for button in sizes_buttons:
                if button.text in ['US M', 'US W', 'US']:
                    button.click()
                    time.sleep(0.1)
                    wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, '#menu-list-pdp-size-selector > div.css-1kgaafq > div')))
                    price_block = driver.find_element(By.CSS_SELECTOR, '#menu-list-pdp-size-selector > div.css-1kgaafq > div')
                    prices = price_block.find_elements(By.TAG_NAME, 'button')
                    for price in prices:
                        size, pr = price.text.split('\n')
                        if size in sizes:
                            worksheet.cell(row=row, column=sizes.index(size)+13).value = pr
                        else:
                            print(f"size not found! new size: {size}")
                            length = len(worksheet[1])
                            worksheet.cell(row=1, column=length+1).value = size
                            sizes.append(size)
                            worksheet.cell(row=row, column=sizes.index(size)+13).value = pr
        except:
            pass
        finally:
            if not flag:
                wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, '#menu-list-pdp-size-selector > div.css-1kgaafq > div')))
                price_block = driver.find_element(By.CSS_SELECTOR, '#menu-list-pdp-size-selector > div.css-1kgaafq > div')
                prices = price_block.find_elements(By.TAG_NAME, 'button')
                for price in prices:
                    size, pr = price.text.split('\n')
                    if size in sizes:
                        worksheet.cell(row=row, column=sizes.index(size) + 13).value = pr
                    else:
                        print(f"size not found! new size: {size}")
                        length = len(worksheet[1])
                        worksheet.cell(row=1, column=length + 1).value = size
                        sizes.append(size)
                        worksheet.cell(row=row, column=sizes.index(size) + 13).value = pr
        time.sleep(random.random())
except Exception as _ex:
    print(_ex)
finally:
    workbook.save('out.xlsx')
    driver.close()
    driver.quit()
